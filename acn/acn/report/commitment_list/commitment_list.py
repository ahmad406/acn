# Report: COMMITMENT LIST
# File: commitment_list.py

import frappe
from frappe import _
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import json


def execute(filters=None):
	columns = get_columns()
	data = get_data(filters)

	total_commitment = len(data)

	total_delivered = len([
		d for d in data
		if d.get("delivery_on_commitment") == 1
	])

	delivery_performance = round(
		(total_delivered / total_commitment) * 100, 2
	) if total_commitment else 0

	message = f"""
	<div style="padding: 12px 0px; line-height: 2; font-size: 14px;">
		<b>TOTAL COMMITMENT:</b> {total_commitment}<br>

		<b>TOTAL DELIVERED ON COMMITMENT:</b>
		{total_delivered}<br>

		<b>DELIVERY PERFORMANCE IN %:</b>
		{delivery_performance}%
	</div>
	"""

	return columns, data, message


def get_columns():
	return [
		{
			"label": _("Commitment Date"),
			"fieldname": "commitment_date",
			"fieldtype": "Date",
			"width": 160,
		},
		{
			"label": _("Customer Name"),
			"fieldname": "customer_name",
			"fieldtype": "Link",
			"options": "Customer",
			"width": 300,
		},
		{
			"label": _("JC NO."),
			"fieldname": "job_card",
			"fieldtype": "Link",
			"options": "Job Card",
			"width": 200,
		},
		{
			"label": _("Part Number"),
			"fieldname": "part_no",
			"fieldtype": "Data",
			"width": 200,
		},
		{
			"label": _("Process"),
			"fieldname": "process_name",
			"fieldtype": "Data",
			"width": 300,
		},
		{
			"label": _("QTY (Nos)"),
			"fieldname": "qty_nos",
			"fieldtype": "Float",
			"width": 140,
		},
		{
			"label": _("QTY (Kgs)"),
			"fieldname": "qty_kgs",
			"fieldtype": "Float",
			"width": 140,
		},
		{
			"label": _("TC DATE"),
			"fieldname": "tc_date",
			"fieldtype": "Date",
			"width": 140,
		},
		{
			"label": _("QTY READY (Nos)"),
			"fieldname": "qty_ready_nos",
			"fieldtype": "Float",
			"width": 160,
		},
		{
			"label": _("QTY READY (Kgs)"),
			"fieldname": "qty_ready_kgs",
			"fieldtype": "Float",
			"width": 160,
		},
		{
			"label": _("Delivery on Commitment"),
			"fieldname": "delivery_on_commitment",
			"fieldtype": "Check",
			"width": 250,
		},
	]


def get_data(filters):
	conditions = ""

	if filters.get("from_date") and filters.get("to_date"):
		conditions += f"""
			AND cdc.commitment_date
			BETWEEN '{filters.get("from_date")}'
			AND '{filters.get("to_date")}'
		"""

	data = frappe.db.sql(
		f"""
		SELECT
			cdc.commitment_date as commitment_date,

			cd.customer_name as customer_name,

			jcp.name as job_card,

			cdc.part_no as part_no,

			jcp.process_name as process_name,

			cdc.qty_nos as qty_nos,

			cdc.qty_kgs as qty_kgs,

			tce.date as tc_date,

			tce.accepted_qty_in_nos as qty_ready_nos,

			tce.accepted_qty_in_kgs as qty_ready_kgs,

			CASE
				WHEN tce.date <= cdc.commitment_date
				THEN 1
				ELSE 0
			END as delivery_on_commitment

		FROM
			`tabCustomer DC child` cdc

		LEFT JOIN `tabCustomer DC` cd
			ON cd.name = cdc.parent

		LEFT JOIN `tabJob Card for process` jcp
			ON jcp.customer_dc = cd.name

		LEFT JOIN `tabTest Certificate entry` tce
			ON tce.customer_dc_id = cd.name

		WHERE
			cd.docstatus = 1
			{conditions}

		ORDER BY
			cdc.commitment_date ASC
		""",
		as_dict=True,
	)

	return data



@frappe.whitelist()
def export_with_summary(filters=None):

	filters = frappe.parse_json(filters)

	columns = get_columns()
	data = get_data(filters)

	total_commitment = len(data)

	total_delivered = len([
		d for d in data
		if d.get("delivery_on_commitment") == 1
	])

	delivery_performance = round(
		(total_delivered / total_commitment) * 100, 2
	) if total_commitment else 0

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Commitment List"

	# Header
	for col_idx, col in enumerate(columns, start=1):

		cell = ws.cell(
			row=1,
			column=col_idx,
			value=col["label"]
		)

		cell.font = Font(bold=True)

		cell.fill = PatternFill(
			"solid",
			fgColor="D9E1F2"
		)

		cell.alignment = Alignment(
			horizontal="center",
			wrap_text=True
		)

	# Data
	for row_idx, row in enumerate(data, start=2):

		for col_idx, col in enumerate(columns, start=1):

			ws.cell(
				row=row_idx,
				column=col_idx,
				value=row.get(col["fieldname"])
			)

	# Summary
	summary_row = len(data) + 4

	ws.cell(
		row=summary_row,
		column=1,
		value="SUMMARY"
	).font = Font(bold=True, size=12)

	summary_row += 1

	summary_data = [
		("TOTAL COMMITMENT", total_commitment),
		("TOTAL DELIVERED ON COMMITMENT", total_delivered),
		("DELIVERY PERFORMANCE IN %", f"{delivery_performance}%"),
	]

	for label, value in summary_data:

		label_cell = ws.cell(
			row=summary_row,
			column=1,
			value=label
		)

		label_cell.font = Font(bold=True)

		label_cell.fill = PatternFill(
			"solid",
			fgColor="F2F2F2"
		)

		value_cell = ws.cell(
			row=summary_row,
			column=2,
			value=value
		)

		value_cell.alignment = Alignment(
			horizontal="right"
		)

		summary_row += 1

	# Auto width
	for col_cells in ws.columns:

		col_letter = col_cells[0].column_letter

		max_len = max(
			(len(str(c.value)) for c in col_cells if c.value),
			default=10
		)

		ws.column_dimensions[col_letter].width = min(
			max_len + 4,
			60
		)

	output = BytesIO()

	wb.save(output)

	output.seek(0)

	frappe.response["type"] = "binary"

	frappe.response["filename"] = "Commitment_List.xlsx"

	frappe.response["filecontent"] = output.getvalue()

	frappe.response["display_content_as"] = "attachment"