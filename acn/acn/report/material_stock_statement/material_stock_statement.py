import frappe
from frappe.utils import flt
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO


def execute(filters=None):
    columns = get_columns()
    data = []

    data.extend(get_opening_stock(filters))
    data.extend(get_material_received(filters))

    return columns, data


def get_columns():
    return [

        {"label": "Particulars", "fieldname": "particulars",
         "fieldtype": "Data", "width": 180},

        {"label": "MRN Date", "fieldname": "mrn_date",
         "fieldtype": "Date", "width": 110},

        {"label": "MRN No", "fieldname": "mrn_no",
         "fieldtype": "Link", "options": "Customer DC", "width": 140},

        {"label": "Party DC No", "fieldname": "party_dc_no",
         "fieldtype": "Data", "width": 130},

        {"label": "Process", "fieldname": "process",
         "fieldtype": "Data", "width": 140},

        {"label": "Item", "fieldname": "item",
         "fieldtype": "Data", "width": 160},

        {"label": "Part Number", "fieldname": "part_number",
         "fieldtype": "Data", "width": 140},

        {"label": "Qty (Nos)", "fieldname": "opening_qty_nos",
         "fieldtype": "Float", "width": 120},

        {"label": "Qty (Kgs)", "fieldname": "opening_qty_kgs",
         "fieldtype": "Float", "width": 120},

        {"label": "Dispatch Qty (Nos)", "fieldname": "dispatch_qty_nos",
         "fieldtype": "Float", "width": 150},

        {"label": "Dispatch Qty (Kgs)", "fieldname": "dispatch_qty_kgs",
         "fieldtype": "Float", "width": 150},

        {"label": "Invoice Nos", "fieldname": "invoice_no",
         "fieldtype": "Data", "width": 250},

        {"label": "Invoice Detail","fieldname": "invoice_detail",
        "fieldtype": "Data","width": 500},

        {"label": "Pending Qty (Nos)", "fieldname": "pending_qty_nos",
         "fieldtype": "Float", "width": 140},

        {"label": "Pending Qty (Kgs)", "fieldname": "pending_qty_kgs",
         "fieldtype": "Float", "width": 140},

        {"label": "Physical Stock (Nos)", "fieldname": "physical_stock_nos",
         "fieldtype": "Float", "width": 160},

        {"label": "Physical Stock (Kgs)", "fieldname": "physical_stock_kgs",
         "fieldtype": "Float", "width": 160},
    ]


# OPENING STOCK

def get_opening_stock(filters):

    rows = []

    opening_data = frappe.db.sql("""
        SELECT
            dc.name AS mrn_no,
            dc.tran_date AS mrn_date,
            child.customer_dc_no,
            child.process_name,
            child.item_name,
            child.part_no,
            child.qty_nos,
            child.qty_kgs,
            dc.sales_order_no
        FROM `tabCustomer DC` dc
        JOIN `tabCustomer DC child` child
            ON child.parent = dc.name
        WHERE dc.docstatus = 1
        AND dc.customer = %(customer)s
        AND dc.tran_date < %(from_date)s
        AND (child.qty_nos > 0 OR child.qty_kgs > 0)
    """, filters, as_dict=True)

    for d in opening_data:

        dispatch = get_dispatch_details(d.mrn_no, d.part_no)

        opening_nos = flt(d.qty_nos)
        opening_kgs = flt(d.qty_kgs)

        rows.append({
            "particulars": "OPENING STOCK",

            "mrn_date": d.mrn_date,
            "mrn_no": d.mrn_no,
            "party_dc_no": d.customer_dc_no,
            "process": d.process_name,
            "item": d.item_name,
            "part_number": d.part_no,

            "opening_qty_nos": opening_nos,
            "opening_qty_kgs": opening_kgs,

            "dispatch_qty_nos": dispatch["nos"],
            "dispatch_qty_kgs": dispatch["kgs"],

            "invoice_no": dispatch["invoice"],
            "invoice_detail": dispatch["detail"],

            "pending_qty_nos": opening_nos - dispatch["nos"],
            "pending_qty_kgs": opening_kgs - dispatch["kgs"],

            "physical_stock_nos": opening_nos - dispatch["nos"],
            "physical_stock_kgs": opening_kgs - dispatch["kgs"],
        })

    return rows


# MATERIAL RECEIVED

def get_material_received(filters):

    rows = []

    received_data = frappe.db.sql("""
        SELECT
            dc.name AS mrn_no,
            dc.tran_date AS mrn_date,
            child.customer_dc_no,
            child.process_name,
            child.item_name,
            child.part_no,
            child.qty_nos,
            child.qty_kgs,
            dc.sales_order_no
        FROM `tabCustomer DC` dc
        JOIN `tabCustomer DC child` child
            ON child.parent = dc.name
        WHERE dc.docstatus = 1
        AND dc.customer = %(customer)s
        AND dc.tran_date BETWEEN %(from_date)s AND %(to_date)s
    """, filters, as_dict=True)

    for d in received_data:

        dispatch = get_dispatch_details(d.mrn_no, d.part_no)

        received_nos = flt(d.qty_nos)
        received_kgs = flt(d.qty_kgs)

        rows.append({
            "particulars": "MATERIAL RECEIVED",

            "mrn_date": d.mrn_date,
            "mrn_no": d.mrn_no,
            "party_dc_no": d.customer_dc_no,
            "process": d.process_name,
            "item": d.item_name,
            "part_number": d.part_no,

            "opening_qty_nos": received_nos,
            "opening_qty_kgs": received_kgs,

            "dispatch_qty_nos": dispatch["nos"],
            "dispatch_qty_kgs": dispatch["kgs"],

            "invoice_no": dispatch["invoice"],
            "invoice_detail": dispatch["detail"],

            "pending_qty_nos": received_nos - dispatch["nos"],
            "pending_qty_kgs": received_kgs - dispatch["kgs"],

            "physical_stock_nos": received_nos - dispatch["nos"],
            "physical_stock_kgs": received_kgs - dispatch["kgs"],
        })

    return rows


def get_dispatch_details(customer_dc, part_no):

    if not customer_dc or not part_no:
        return {"nos": 0, "kgs": 0, "invoice": "", "detail": ""}

    result = frappe.db.sql("""
        SELECT
            SUM(sii.d_qty_in_nos) AS nos,
            SUM(sii.d_qty_in_kgs) AS kgs,

            GROUP_CONCAT(
                DISTINCT CONCAT(
                    '(',
                    si.name,
                    ' - ',
                    FORMAT(IFNULL(sii.d_qty_in_nos,0), 2),
                    ' Nos - ',
                    DATE_FORMAT(si.posting_date,'%%d-%%m-%%Y'),
                    ')'
                )
                SEPARATOR ', '
            ) AS invoice_detail,

            GROUP_CONCAT(DISTINCT si.name) AS invoices

        FROM `tabSales Invoice` si
        JOIN `tabSales Invoice Item` sii
            ON sii.parent = si.name
        WHERE si.docstatus = 1
        AND sii.customer_dc_id = %s
        AND sii.part_no = %s
        """, (customer_dc, part_no), as_dict=True)

    if result and result[0]:
        r = result[0]
        return {
            "nos": flt(r.nos),
            "kgs": flt(r.kgs),
            "invoice": r.invoices or "",
            "detail": r.invoice_detail or ""
        }

    return {"nos": 0, "kgs": 0, "invoice": "", "detail": ""}



@frappe.whitelist()
def export_with_summary(filters=None):

    filters = frappe.parse_json(filters)

    # Get report data
    result = execute(filters)
    columns = result[0]
    data = result[1]

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Material Stock Statement"

    # -----------------------------
    # WRITE TABLE HEADERS
    # -----------------------------
    col_index = 1
    for col in columns:
        ws.cell(row=1, column=col_index).value = col.get("label")
        ws.cell(row=1, column=col_index).font = Font(bold=True)
        col_index += 1

    # -----------------------------
    # WRITE DATA ROWS
    # -----------------------------
    row_index = 2
    for row in data:
        col_index = 1
        for col in columns:
            ws.cell(row=row_index, column=col_index).value = row.get(col.get("fieldname"))
            col_index += 1
        row_index += 1

    # -----------------------------
    # CALCULATE TOTALS
    # -----------------------------
    opening_total = sum(
        d.get("opening_qty_nos", 0)
        for d in data if d.get("particulars") == "OPENING STOCK"
    )

    received_total = sum(
        d.get("opening_qty_nos", 0)
        for d in data if d.get("particulars") == "MATERIAL RECEIVED"
    )

    dispatched_total = sum(d.get("dispatch_qty_nos", 0) for d in data)
    closing_total = sum(d.get("pending_qty_nos", 0) for d in data)

    row_index += 3

    # -----------------------------
    # LEFT SIDE TERMS
    # -----------------------------
    terms_col = 1

    ws.merge_cells(start_row=row_index, start_column=terms_col,
                   end_row=row_index, end_column=5)

    cell = ws.cell(row=row_index, column=terms_col)
    cell.value = "If this statement is related to your material laying with us, please note the following:"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(wrap_text=True)

    row_index += 1
    ws.merge_cells(start_row=row_index, start_column=terms_col,
                   end_row=row_index, end_column=5)
    ws.cell(row=row_index, column=terms_col).value = \
        "1. Review the details for accuracy and ensure they align with your stock book/ledger."
    ws.cell(row=row_index, column=terms_col).alignment = Alignment(wrap_text=True)

    row_index += 1
    ws.merge_cells(start_row=row_index, start_column=terms_col,
                   end_row=row_index, end_column=5)
    ws.cell(row=row_index, column=terms_col).value = \
        "2. If any discrepancy in stock quantity or missing items, please revert within one week."
    ws.cell(row=row_index, column=terms_col).alignment = Alignment(wrap_text=True)

    row_index += 2
    ws.merge_cells(start_row=row_index, start_column=terms_col,
                   end_row=row_index, end_column=5)

    cell = ws.cell(row=row_index, column=terms_col)
    cell.value = ("Kindly sign and return the statement within one week. "
                  "If no reply is received, we will assume the stock is correct and undisputed.")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(wrap_text=True)

    # -----------------------------
    # RIGHT SIDE SUMMARY
    # -----------------------------
    summary_col = 7
    summary_row = row_index - 4
    right_align = Alignment(horizontal="right")

    ws.cell(row=summary_row, column=summary_col).value = "SUMMARY"
    ws.cell(row=summary_row, column=summary_col).font = Font(bold=True)

    summary_row += 1
    ws.cell(row=summary_row, column=summary_col).value = "Opening Stock"
    ws.cell(row=summary_row, column=summary_col + 1).value = opening_total
    ws.cell(row=summary_row, column=summary_col + 1).alignment = right_align

    summary_row += 1
    ws.cell(row=summary_row, column=summary_col).value = "Material Receipt"
    ws.cell(row=summary_row, column=summary_col + 1).value = received_total
    ws.cell(row=summary_row, column=summary_col + 1).alignment = right_align

    summary_row += 1
    ws.cell(row=summary_row, column=summary_col).value = "Material Dispatched"
    ws.cell(row=summary_row, column=summary_col + 1).value = dispatched_total
    ws.cell(row=summary_row, column=summary_col + 1).alignment = right_align

    summary_row += 1
    ws.cell(row=summary_row, column=summary_col).value = "Closing Stock"
    ws.cell(row=summary_row, column=summary_col + 1).value = closing_total
    ws.cell(row=summary_row, column=summary_col + 1).alignment = right_align

    # -----------------------------
    # AUTO COLUMN WIDTH
    # -----------------------------
    for column_cells in ws.columns:
        length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    length = max(length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(length + 3, 60)

    # -----------------------------
    # DIRECT DOWNLOAD
    # -----------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response["type"] = "binary"
    frappe.response["filename"] = "Material_Stock_Statement.xlsx"
    frappe.response["filecontent"] = output.getvalue()
    frappe.response["display_content_as"] = "attachment"