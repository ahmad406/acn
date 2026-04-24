import frappe
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO



def execute(filters=None):

	columns = get_columns()
	data = get_data(filters)	
	total = len(data)
	
	closed = len([r for r in data if (r.get("status") or "").lower() in ("converted", "closed")])
	converted = len([r for r in data if (r.get("status") or "").lower() in ("converted", "won")])
	conversion_pct = round((converted / total * 100), 2) if total else 0

	message = f"""
        <div style="padding: 10px 0; font-size: 14px; line-height: 2;">
            <b>Total No of Enquiries:</b> {total}<br>
            <b>Total No of Enquiries Closed:</b> {closed}<br>
            <b>Conversion of Enquiries:</b> {converted} ({conversion_pct}%)
        </div>
    """

	return columns, data, message

def get_columns():
	return [
		{"label": "Date of Enquiry", "fieldname": "transaction_date", "fieldtype": "Date",     "width": 120},
		{"label": "Enquiry Type",    "fieldname": "enquiry_type",     "fieldtype": "Data",     "width": 120},
		{"label": "Source",          "fieldname": "source",           "fieldtype": "Data",     "width": 110},
		{"label": "Customer Name",   "fieldname": "customer_name",    "fieldtype": "Link",     "options": "Lead", "width": 160},
		{"label": "Sector",          "fieldname": "industry",         "fieldtype": "Data",     "width": 120},
		{"label": "Territory",       "fieldname": "territory",        "fieldtype": "Data",     "width": 120},
		{"label": "Contact Person",  "fieldname": "contact_person",   "fieldtype": "Data",     "width": 140},
		{"label": "Contact No",      "fieldname": "mobile_no",        "fieldtype": "Data",     "width": 120},
		{"label": "E-Mail ID",       "fieldname": "email_id",         "fieldtype": "Data",     "width": 160},
		{"label": "Process",         "fieldname": "process",          "fieldtype": "Data",     "width": 180},
		{"label": "Quotation",       "fieldname": "quotation_name",   "fieldtype": "Link",     "options": "Quotation", "width": 140},
		{"label": "Rates Quoted",    "fieldname": "rates_quoted",     "fieldtype": "Currency", "width": 130},
		{"label": "Total Value",     "fieldname": "total_value",      "fieldtype": "Currency", "width": 130},
		{"label": "Enq Status",      "fieldname": "status",           "fieldtype": "Data",     "width": 120},
		{"label": "Remarks",         "fieldname": "remarks",          "fieldtype": "Data",     "width": 240},
	]


def get_data(filters=None):
	filters = filters or {}

	conditions = ""
	values = []

	if filters.get("from_date"):
		conditions += " AND DATE(l.creation) >= %s"
		values.append(filters["from_date"])
	if filters.get("to_date"):
		conditions += " AND DATE(l.creation) <= %s"
		values.append(filters["to_date"])

	leads = frappe.db.sql("""
        SELECT
            l.name       AS lead_id,
            l.company_name AS customer_name,
            l.creation   AS transaction_date,
            l.enquiry_type,
            l.source,
            l.industry,
            l.territory,
            l.first_name AS contact_person,
            l.mobile_no,
            l.email_id,
            l.status
        FROM `tabLead` l
        WHERE 1=1 {conditions}
        ORDER BY l.creation
    """.format(conditions=conditions), values, as_dict=True)

	if not leads:
		return []

	lead_ids = [l.lead_id for l in leads]
	ph_leads = ", ".join(["%s"] * len(lead_ids))

	opp_meta_map  = {}
	opp_items_map = {}
	lead_opps_map = {}

	opp_rows = frappe.db.sql("""
		SELECT
			o.name               AS opp_id,
			o.party_name         AS lead_id,
			o.opportunity_amount AS total_value,
			oi.item_name         AS process
		FROM `tabOpportunity` o
		LEFT JOIN `tabOpportunity Item` oi ON oi.parent = o.name
		WHERE o.opportunity_from = 'Lead'
		  AND o.party_name IN ({ph})
		ORDER BY o.name, oi.idx
	""".format(ph=ph_leads), lead_ids, as_dict=True)

	for o in opp_rows:
		if o.opp_id not in opp_meta_map:
			opp_meta_map[o.opp_id] = {
				"lead_id":     o.lead_id,
				"total_value": o.total_value
			}
			lead_opps_map.setdefault(o.lead_id, [])
			if o.opp_id not in lead_opps_map[o.lead_id]:
				lead_opps_map[o.lead_id].append(o.opp_id)
		if o.process:
			opp_items_map.setdefault(o.opp_id, []).append(o.process)

	qtn_map = {}

	opp_ids = list(opp_meta_map.keys())
	if opp_ids:
		ph_opps = ", ".join(["%s"] * len(opp_ids))

		qtn_rows = frappe.db.sql("""
			SELECT
				q.name AS quotation_name,
				q.opportunity AS opp_id,
				qi.rate AS rates_quoted
			FROM `tabQuotation` q
			LEFT JOIN `tabQuotation Item` qi
				ON qi.parent = q.name AND qi.idx = 1
			WHERE q.opportunity IN ({ph})
			  AND q.docstatus != 2
		""".format(ph=ph_opps), opp_ids, as_dict=True)

		for q in qtn_rows:
			qtn_map.setdefault(q.opp_id, []).append(q)

	remarks_map = {}

	if opp_ids:
		ph_opps = ", ".join(["%s"] * len(opp_ids))

		task_rows = frappe.db.sql("""
			SELECT
				reference_name AS opp_id,
				GROUP_CONCAT(description ORDER BY creation SEPARATOR ' | ') AS val
			FROM `tabToDo`
			WHERE reference_type = 'Opportunity'
			  AND reference_name IN ({ph})
			GROUP BY reference_name
		""".format(ph=ph_opps), opp_ids, as_dict=True)

		comment_rows = frappe.db.sql("""
			SELECT
				parent AS opp_id,
				GROUP_CONCAT(note ORDER BY creation SEPARATOR ' | ') AS val
			FROM `tabCRM Note`
			  where parent IN ({ph})
			GROUP BY parent
		""".format(ph=ph_opps), opp_ids, as_dict=True)

		for row in task_rows + comment_rows:
			if row.val:
				remarks_map.setdefault(row.opp_id, []).append(row.val)

	result = []

	for lead in leads:
		opp_ids_for_lead = lead_opps_map.get(lead.lead_id)

		if not opp_ids_for_lead:
			result.append(make_row(lead, None, None, None, ""))
			continue

		for opp_id in opp_ids_for_lead:
			meta    = opp_meta_map[opp_id]
			items   = opp_items_map.get(opp_id) or [None]
			qtns    = qtn_map.get(opp_id) or [None]
			remarks = " | ".join(remarks_map.get(opp_id, []))

			for item in items:
				for qtn in qtns:
					result.append(make_row(lead, meta, item, qtn, remarks))

	return result


def make_row(lead, opp_meta, item, qtn, remarks):
	return {
		"transaction_date": lead.transaction_date,
		"enquiry_type":     lead.enquiry_type,
		"source":           lead.source,
		"customer_name":    lead.company_name,
		"industry":         lead.industry,
		"territory":        lead.territory,
		"contact_person":   lead.contact_person,
		"mobile_no":        lead.mobile_no,
		"email_id":         lead.email_id,
		"process":          item or "",
		"quotation_name":   qtn.quotation_name if qtn else "",
		"rates_quoted":     qtn.rates_quoted   if qtn else 0,
		"total_value":      opp_meta["total_value"] if opp_meta else 0,
		"status":           lead.status,
		"remarks":          remarks,
	}

@frappe.whitelist()
def export_with_summary(filters=None):
	filters = frappe.parse_json(filters)
	columns = get_columns()
	data = get_data(filters)

	# Summary calculations
	total = len(data)
	closed = len([r for r in data if (r.get("status") or "").lower() in ("converted", "closed")])
	converted = len([r for r in data if (r.get("status") or "").lower() in ("converted", "won")])
	conversion_pct = round((converted / total * 100), 2) if total else 0

	# Workbook
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Enquiry Register"

	# Header row
	for col_idx, col in enumerate(columns, start=1):
		cell = ws.cell(row=1, column=col_idx, value=col["label"])
		cell.font = Font(bold=True)
		cell.fill = PatternFill("solid", fgColor="D9E1F2")
		cell.alignment = Alignment(horizontal="center", wrap_text=True)

	# Data rows
	for row_idx, row in enumerate(data, start=2):
		for col_idx, col in enumerate(columns, start=1):
			ws.cell(row=row_idx, column=col_idx, value=row.get(col["fieldname"]))

	# Summary block
	summary_row = len(data) + 4

	summary_data = [
		("Total No of Enquiries", total),
		("Total No of Enquiries Closed", closed),
		("Conversion of Enquiries", f"{converted} ({conversion_pct}%)"),
	]

	ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
	summary_row += 1

	for label, value in summary_data:
		label_cell = ws.cell(row=summary_row, column=1, value=label)
		label_cell.font = Font(bold=True)
		label_cell.fill = PatternFill("solid", fgColor="F2F2F2")

		value_cell = ws.cell(row=summary_row, column=2, value=value)
		value_cell.alignment = Alignment(horizontal="right")

		summary_row += 1

	# Auto column width
	for col_cells in ws.columns:
		col_letter = col_cells[0].column_letter
		max_len = max(
			(len(str(c.value)) for c in col_cells if c.value),
			default=10
		)
		ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

	# Stream response
	output = BytesIO()
	wb.save(output)
	output.seek(0)

	frappe.response["type"] = "binary"
	frappe.response["filename"] = "Enquiry_Register.xlsx"
	frappe.response["filecontent"] = output.getvalue()
	frappe.response["display_content_as"] = "attachment"